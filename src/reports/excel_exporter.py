"""
Exportador de reportes ejecutivos a Excel - TORO · Resumen de Cuentas
Autor: Sistema TORO
"""
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from typing import Dict

class ExcelExporter:
    """
    Genera reportes ejecutivos en Excel con múltiples hojas y formato profesional.
    """

    def __init__(self, df: pd.DataFrame, metricas: Dict):
        """
        Args:
            df: DataFrame con movimientos categorizados
            metricas: Diccionario con métricas calculadas
        """
        self.df = df
        self.metricas = metricas

    def _formatear_monto(self, valor: float) -> str:
        """
        Formatea un monto manejando valores NaN.

        Args:
            valor: Valor a formatear

        Returns:
            String formateado
        """
        if pd.isna(valor):
            return "No disponible"
        return f"${valor:,.2f}"

    def exportar(self, ruta_salida: str):
        """
        Exporta el reporte ejecutivo a Excel.

        Args:
            ruta_salida: Ruta del archivo de salida
        """
        print(f"\nGenerando reporte ejecutivo Excel...")

        # Crear Excel writer
        with pd.ExcelWriter(ruta_salida, engine='openpyxl') as writer:
            # Hoja 1: Resumen
            self._crear_hoja_resumen(writer)

            # Hoja 2: Ingresos
            self._crear_hoja_ingresos(writer)

            # Hoja 3: Egresos
            self._crear_hoja_egresos(writer)

            # Hoja 4: Prestadores
            self._crear_hoja_prestadores(writer)

            # Hoja 5: Sin Clasificar
            self._crear_hoja_sin_clasificar(writer)

        # Aplicar formato
        self._aplicar_formato(ruta_salida)

        print(f"OK Reporte ejecutivo generado: {ruta_salida}")

    def _crear_hoja_resumen(self, writer):
        """
        Crea la hoja de resumen ejecutivo.
        """
        datos = [
            ['SANARTE - RESUMEN EJECUTIVO', ''],
            ['', ''],
            ['SALDOS BANCARIOS', ''],
            ['Saldo Inicial', self._formatear_monto(self.metricas['saldo_inicial'])],
            ['Total Ingresos', f"${self.metricas['total_ingresos']:,.2f}"],
            ['Total Egresos', f"${self.metricas['total_egresos']:,.2f}"],
            ['Saldo Final', self._formatear_monto(self.metricas['saldo_final'])],
            ['Variacion del Mes', f"${self.metricas['variacion']:,.2f}"],
            ['', ''],
        ]

        # Agregar advertencia si hay diferencia en validación
        if not self.metricas['validacion_saldos_ok']:
            datos.append(['ADVERTENCIA', ''])
            datos.append(['Diferencia en validacion', f"${self.metricas['diferencia_validacion']:,.2f}"])
            datos.append(['Detalle', 'El saldo final no coincide con Saldo Inicial + Ingresos - Egresos'])
            datos.append(['', ''])

        datos.extend([
            ['CLASIFICACION', ''],
            ['Total Movimientos', self.metricas['total_movimientos']],
            ['Clasificados', self.metricas['movimientos_clasificados']],
            ['Sin Clasificar', self.metricas['movimientos_sin_clasificar']],
            ['% Clasificados', f"{self.metricas['porcentaje_clasificado']:.1f}%"],
            ['', ''],
            ['DESGLOSE INGRESOS', 'Monto'],
        ])

        # Agregar ingresos por subcategoría (ordenados de mayor a menor)
        ingresos_ordenados = sorted(self.metricas['ingresos_por_subcategoria'].items(),
                                    key=lambda x: x[1], reverse=True)
        for sub, monto in ingresos_ordenados:
            datos.append([sub, f"${monto:,.2f}"])

        datos.append(['', ''])
        datos.append(['DESGLOSE EGRESOS', 'Monto'])

        # Agregar egresos por subcategoría (ordenados de mayor a menor)
        egresos_ordenados = sorted(self.metricas['egresos_por_subcategoria'].items(),
                                   key=lambda x: x[1], reverse=True)
        for sub, monto in egresos_ordenados:
            datos.append([sub, f"${monto:,.2f}"])

        # Crear DataFrame y exportar
        df_resumen = pd.DataFrame(datos, columns=['Concepto', 'Valor'])
        df_resumen.to_excel(writer, sheet_name='Resumen', index=False)

    def _crear_hoja_ingresos(self, writer):
        """
        Crea la hoja de desglose de ingresos con resumen de saldos.
        """
        df_ingresos = self.df[self.df['Tipo_Movimiento'] == 'Ingreso'].copy()

        if len(df_ingresos) == 0:
            # Hoja vacía con mensaje
            pd.DataFrame(['No hay ingresos registrados']).to_excel(writer, sheet_name='Ingresos', index=False, header=False)
            return

        # Crear hoja con resumen y detalle
        # Primero agregamos el resumen de saldos
        datos_resumen = [
            ['ANALISIS DE INGRESOS - RESUMEN', ''],
            ['', ''],
            ['Saldo Inicial', self._formatear_monto(self.metricas['saldo_inicial'])],
            ['Total Ingresos', f"${self.metricas['total_ingresos']:,.2f}"],
            ['Total Egresos', f"${self.metricas['total_egresos']:,.2f}"],
            ['Saldo Final', self._formatear_monto(self.metricas['saldo_final'])],
            ['', ''],
            ['DESGLOSE POR SUBCATEGORIA', 'Monto'],
        ]

        # Agregar ingresos por subcategoría (ordenados de mayor a menor)
        ingresos_ordenados = sorted(self.metricas['ingresos_por_subcategoria'].items(),
                                    key=lambda x: x[1], reverse=True)
        for sub, monto in ingresos_ordenados:
            porcentaje = (monto / self.metricas['total_ingresos'] * 100) if self.metricas['total_ingresos'] > 0 else 0
            datos_resumen.append([sub, f"${monto:,.2f} ({porcentaje:.1f}%)"])

        datos_resumen.extend([
            ['', ''],
            ['', ''],
            ['DETALLE DE MOVIMIENTOS', ''],
        ])

        # Crear DataFrame de resumen
        df_resumen = pd.DataFrame(datos_resumen, columns=['Concepto', 'Valor'])

        # Seleccionar columnas relevantes
        columnas = ['Fecha', 'Concepto', 'Detalle', 'Crédito', 'Categoria_Final',
                   'Persona_Nombre', 'Es_DEBIN', 'Banco']

        df_export = df_ingresos[columnas].copy()

        # Ordenar por fecha descendente
        df_export = df_export.sort_values('Fecha', ascending=False)

        # Crear archivo temporal con ambas secciones
        # Primero el resumen
        df_resumen.to_excel(writer, sheet_name='Ingresos', index=False, startrow=0)

        # Luego el detalle (después del resumen + una fila de separación)
        start_row = len(df_resumen) + 2
        df_export.to_excel(writer, sheet_name='Ingresos', index=False, startrow=start_row)

    def _crear_hoja_egresos(self, writer):
        """
        Crea la hoja de desglose de egresos con resumen de saldos.
        """
        df_egresos = self.df[self.df['Tipo_Movimiento'] == 'Egreso'].copy()

        if len(df_egresos) == 0:
            pd.DataFrame(['No hay egresos registrados']).to_excel(writer, sheet_name='Egresos', index=False, header=False)
            return

        # Crear hoja con resumen y detalle
        # Primero agregamos el resumen de saldos
        datos_resumen = [
            ['ANALISIS DE GASTOS - RESUMEN', ''],
            ['', ''],
            ['Saldo Inicial', self._formatear_monto(self.metricas['saldo_inicial'])],
            ['Total Egresos', f"${self.metricas['total_egresos']:,.2f}"],
            ['Total Ingresos', f"${self.metricas['total_ingresos']:,.2f}"],
            ['Saldo Final', self._formatear_monto(self.metricas['saldo_final'])],
            ['', ''],
            ['DESGLOSE POR SUBCATEGORIA', 'Monto'],
        ]

        # Agregar egresos por subcategoría (ordenados de mayor a menor)
        egresos_ordenados = sorted(self.metricas['egresos_por_subcategoria'].items(),
                                   key=lambda x: x[1], reverse=True)
        for sub, monto in egresos_ordenados:
            porcentaje = (monto / self.metricas['total_egresos'] * 100) if self.metricas['total_egresos'] > 0 else 0
            datos_resumen.append([sub, f"${monto:,.2f} ({porcentaje:.1f}%)"])

        datos_resumen.extend([
            ['', ''],
            ['', ''],
            ['DETALLE DE MOVIMIENTOS', ''],
        ])

        # Crear DataFrame de resumen
        df_resumen = pd.DataFrame(datos_resumen, columns=['Concepto', 'Valor'])

        # Seleccionar columnas relevantes de egresos
        columnas = ['Fecha', 'Concepto', 'Detalle', 'Débito', 'Categoria_Final',
                   'Persona_Nombre', 'Banco']

        df_export = df_egresos[columnas].copy()

        # Ordenar por fecha descendente
        df_export = df_export.sort_values('Fecha', ascending=False)

        # Crear archivo temporal con ambas secciones
        # Primero el resumen
        df_resumen.to_excel(writer, sheet_name='Egresos', index=False, startrow=0)

        # Luego el detalle (después del resumen + una fila de separación)
        start_row = len(df_resumen) + 2
        df_export.to_excel(writer, sheet_name='Egresos', index=False, startrow=start_row)

    def _crear_hoja_prestadores(self, writer):
        """
        Crea la hoja de Top 15 Egresos (todos los egresos, no solo prestadores).
        """
        # Filtrar todos los egresos
        df_egresos = self.df[self.df['Tipo_Movimiento'] == 'Egreso'].copy()

        if len(df_egresos) == 0:
            pd.DataFrame(['No hay egresos registrados']).to_excel(writer, sheet_name='Top Egresos', index=False, header=False)
            return

        # Calcular resumen
        total_egresos = df_egresos['Débito'].sum()
        cantidad_egresos = len(df_egresos)
        promedio_egreso = total_egresos / cantidad_egresos if cantidad_egresos > 0 else 0

        # Crear resumen
        datos_resumen = [
            ['TOP 15 EGRESOS MÁS GRANDES', ''],
            ['', ''],
            ['RESUMEN DE EGRESOS', ''],
            ['Total de Egresos', f"${total_egresos:,.2f}"],
            ['Cantidad de Egresos', f"{cantidad_egresos:,}"],
            ['Promedio por Egreso', f"${promedio_egreso:,.2f}"],
            ['', ''],
            ['', ''],
            ['DETALLE - TOP 15 MAYORES EGRESOS', ''],
        ]

        df_resumen = pd.DataFrame(datos_resumen, columns=['Concepto', 'Valor'])

        # Ordenar egresos por monto (mayor a menor) y tomar top 15
        df_top = df_egresos.nlargest(15, 'Débito').copy()

        # Agregar ranking
        df_top.insert(0, 'Ranking', range(1, len(df_top) + 1))

        # Seleccionar columnas relevantes
        columnas = ['Ranking', 'Fecha', 'Concepto', 'Detalle', 'Categoria_Final', 'Débito']
        df_export = df_top[columnas].copy()

        # Renombrar columna Débito a Monto para mayor claridad
        df_export = df_export.rename(columns={'Débito': 'Monto'})

        # Exportar resumen primero
        df_resumen.to_excel(writer, sheet_name='Top Egresos', index=False, startrow=0)

        # Luego exportar el detalle
        start_row = len(df_resumen) + 2
        df_export.to_excel(writer, sheet_name='Top Egresos', index=False, startrow=start_row)

    def _crear_hoja_sin_clasificar(self, writer):
        """
        Crea la hoja de movimientos sin clasificar.
        """
        df_sin_clasificar = self.df[self.df['Categoria_Principal'] == 'Sin Clasificar'].copy()

        if len(df_sin_clasificar) == 0:
            pd.DataFrame(['Todos los movimientos estan clasificados']).to_excel(writer, sheet_name='Sin Clasificar', index=False, header=False)
            return

        # Seleccionar columnas relevantes
        columnas = ['Fecha', 'Concepto', 'Detalle', 'Débito', 'Crédito', 'Banco']

        df_export = df_sin_clasificar[columnas].copy()

        # Ordenar por fecha descendente
        df_export = df_export.sort_values('Fecha', ascending=False)

        # Exportar
        df_export.to_excel(writer, sheet_name='Sin Clasificar', index=False)

    def _aplicar_formato(self, ruta_archivo: str):
        """
        Aplica formato profesional al Excel.

        Args:
            ruta_archivo: Ruta del archivo Excel
        """
        # Cargar el workbook
        wb = load_workbook(ruta_archivo)

        # Estilos
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        title_font = Font(bold=True, size=14)
        title_alignment = Alignment(horizontal='left', vertical='center')

        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Formatear cada hoja
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Auto-ajustar anchos de columna
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Aplicar formato a headers (fila 1)
            if ws.max_row > 0:
                for cell in ws[1]:
                    if cell.value:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = border

            # Si es la hoja de Resumen, Ingresos, Egresos o Top Egresos, aplicar formato especial
            if sheet_name in ['Resumen', 'Ingresos', 'Egresos', 'Top Egresos']:
                ws['A1'].font = Font(bold=True, size=16, color='4472C4')
                ws['A1'].alignment = title_alignment

                # Hacer títulos de sección en negrita
                for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
                    if row[0].value and isinstance(row[0].value, str):
                        if row[0].value.isupper() or 'TOTAL' in row[0].value.upper():
                            row[0].font = Font(bold=True, size=11)

        # Guardar cambios
        wb.save(ruta_archivo)
